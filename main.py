#!/usr/bin/env python
import pandas as pd
import numpy as np
from pathlib import Path
from openpyxl import load_workbook
from openpyxl import Workbook
from copy import copy
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import colors, Font
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule
from openpyxl.utils import get_column_letter

def unmerge(*args):
    for arg in args:
        for sheet in arg:
            ws = sheet
            print('Now in sheet: ' + ws.title)
            # Remove merged cells
            mergedRanges=ws.merged_cells.ranges
            o=0
            while mergedRanges:
                for entry in mergedRanges:
                    o=+1
                    print("  unMerging: " + str(o) + ": " +str(entry))
                    ws.unmerge_cells(str(entry))

def diff_pd(df_old, df_new):

    """Identify differences between two pandas DataFrames"""

    # Take copy
    df_diff = df_new.copy()

    # Get columns
    cols_old = df_old.columns
    cols_new = df_new.columns
    sharedCols = list(set(cols_old).intersection(cols_new))

    # Determine differences
    for row in df_diff.index:
        for col in cols_new:
            value_new = df_new.loc[row, col]
            # Differences in shared col and rows
            if (row in df_old.index) and (row in df_new.index):
                if col in sharedCols:
                    value_old = df_old.loc[row, col]
                    if (value_old == value_new) | (
                        str(value_old) == "nan" and str(value_new) == "nan"
                    ):
                        df_diff.loc[row, col] = df_new.loc[row, col]
                    elif str(value_old) == "nan" and str(value_new) != "nan":
                        df_diff.loc[row, col] = ("Added:{}").format(value_new)
                    elif str(value_new) == "nan" and str(value_old) != "nan":
                        df_diff.loc[row, col] = ("Removed:{}").format(value_old)
                    else:
                        df_diff.loc[row, col] = ("{}→{}").format(value_old, value_new)
                else:
                    if str(value_new) != "nan":
                        df_diff.loc[row, col] = ("Added:{}").format(value_new)
            else:
                # Differences for rows only in new file
                if str(value_new) != "nan":
                    df_diff.loc[row, col] = ("Added:{}").format(value_new)

    for row in df_old.index:
        if row not in df_new.index:
            # Differences in rows only in old file
            for col in cols_old:
                value_old = df_old.loc[row, col]
                if str(value_old) == "nan":
                    df_diff.loc[row, col] = np.nan
                else:
                    df_diff.loc[row, col] = ("Removed:{}").format(value_old)
        # Differences in cols only in old file
        for col in cols_old:
            if col not in cols_new:
                value_old = df_old.loc[row, col]
                if str(value_old) == "nan":
                    df_diff.loc[row, col] = np.nan
                else:
                    df_diff.loc[row, col] = ("Removed:{}").format(value_old)

    df_diff = df_diff.sort_index().fillna("")

    return df_diff


def compare_excel(path_new, path_old,file_path, **kwargs):

    # Get Conditional formating styles
    remove_red_font = Font(color="00FF0000")
    add_green_font = Font(color="00008000")
    diff_orange_font = Font(color="00FF6600")
    diff_style1 = DifferentialStyle(font=diff_orange_font)
    rule1 = Rule(type="expression", dxf=diff_style1)
    rule1.formula = ['NOT(ISERROR(SEARCH("→",A1)))']
    diff_style2 = DifferentialStyle(font=remove_red_font)
    rule2 = Rule(type="expression", dxf=diff_style2)
    rule2.formula = ['NOT(ISERROR(SEARCH("Removed",A1)))']
    diff_style3 = DifferentialStyle(font=add_green_font)
    rule3 = Rule(type="expression", dxf=diff_style3)
    rule3.formula = ['NOT(ISERROR(SEARCH("Added",A1)))']

    # Save output and format
    fname = "{} vs {}.xlsx".format(path_new.stem, path_old.stem)

    print("Comparison file '{}' created.".format(fname))

    read_from_new = load_workbook(path_new, data_only=True)
    read_from_old = load_workbook(path_old, data_only=True)
    write_to = Workbook()

    #######################################
    #Unmerge cells so comparison can occur
    unmerge(read_from_new,read_from_old)
    #######################################

    # Shared worksheets
    sheets_new = read_from_new.worksheets
    sheets_old = read_from_old.worksheets

    sheet_names_new = read_from_new.sheetnames
    sheet_names_old = read_from_old.sheetnames

    shared_sheet_names = list(set(sheet_names_new).intersection(sheet_names_old))

    for sheet in read_from_new:
        # Read in new excel file
        sheet_name = sheet.title

        new_df = pd.read_excel(path_new, sheet_name=sheet_name, **kwargs)
        new_df_as_rows = dataframe_to_rows(new_df, index=False)

        if sheet_name in shared_sheet_names:

            old_df = pd.read_excel(path_old, sheet_name=sheet_name, **kwargs)
            # Get difference in shared worksheets
            df_diff = diff_pd(old_df, new_df)

            # Use openpyxl to convert dataframes.
            df_diff_as_rows = dataframe_to_rows(df_diff, index=False)
            if len(sheet_name) > 20:
                sheet_name = sheet_name[:20]
                print("Warning:some sheet names in comparison file truncated")

            write_to.create_sheet("Diff-{}".format(sheet_name))
            write_sheet = write_to["Diff-{}".format(sheet_name)]
            for r_idx, row in enumerate(df_diff_as_rows, 1):
                for c_idx, value in enumerate(row, 1):
                    if "Unnamed" in str(value):
                        value = ""
                    else:
                        value
                    cell = sheet.cell(row=r_idx, column=c_idx, value=value)
                    diff_cell = write_sheet.cell(row=r_idx, column=c_idx, value=value)
                    if cell.has_style:
                        diff_cell.font = copy(cell.font)
                        diff_cell.border = copy(cell.border)
                        diff_cell.fill = copy(cell.fill)
                        diff_cell.number_format = copy(cell.number_format)
                        diff_cell.protection = copy(cell.protection)
                        diff_cell.alignment = copy(cell.alignment)
            col_letter = get_column_letter(c_idx)
            write_sheet.conditional_formatting.add(
                "$A1:$" + col_letter + str(r_idx), rule1
            )
            write_sheet.conditional_formatting.add(
                "$A1:$" + col_letter + str(r_idx), rule2
            )
            write_sheet.conditional_formatting.add(
                "$A1:$" + col_letter + str(r_idx), rule3
            )
        else:
            if len(sheet_name) > 20:
                sheet_name = sheet_name[:20]
                print("Warning:some sheet names in comparison file truncated")
            write_to.create_sheet("New-{}".format(sheet_name))
            write_sheet = write_to["New-{}".format(sheet_name)]
            for r_idx, row in enumerate(new_df_as_rows, 1):
                for c_idx, value in enumerate(row, 1):
                    if "Unnamed" in str(value):
                        value = ""
                    else:
                        value
                    cell = sheet.cell(row=r_idx, column=c_idx, value=value)
                    diff_cell = write_sheet.cell(row=r_idx, column=c_idx, value=value)
                    diff_cell.font = add_green_font
                    if cell.has_style:
                        diff_cell.border = copy(cell.border)
                        diff_cell.fill = copy(cell.fill)
                        diff_cell.number_format = copy(cell.number_format)
                        diff_cell.protection = copy(cell.protection)
                        diff_cell.alignment = copy(cell.alignment)

    for sheet in read_from_old:
        sheet_name = sheet.title
        if sheet_name not in shared_sheet_names:
            old_df = pd.read_excel(path_old, sheet_name=sheet_name, **kwargs)
            old_df_as_rows = dataframe_to_rows(old_df, index=False)
            if len(sheet_name) > 20:
                sheet_name = sheet_name[:20]
                print("Warning:some sheet names in comparison file truncated")
            write_to.create_sheet("Old-{}".format(sheet_name))
            write_sheet = write_to["Old-{}".format(sheet_name)]
            for r_idx, row in enumerate(old_df_as_rows, 1):
                for c_idx, value in enumerate(row, 1):
                    if "Unnamed" in str(value):
                        value = ""
                    else:
                        value
                    cell = sheet.cell(row=r_idx, column=c_idx, value=value)
                    diff_cell = write_sheet.cell(row=r_idx, column=c_idx, value=value)
                    diff_cell.font = remove_red_font
                    if cell.has_style:
                        diff_cell.border = copy(cell.border)
                        diff_cell.fill = copy(cell.fill)
                        diff_cell.number_format = copy(cell.number_format)
                        diff_cell.protection = copy(cell.protection)
                        diff_cell.alignment = copy(cell.alignment)
    # Remove default sheet
    if "Sheet" in write_to.sheetnames:
        sheet_index = write_to["Sheet"]
        write_to.remove(sheet_index)

    # Save and Close
    print(file_path)
    write_to.save(file_path)
    write_to.close()


def main_gui(path1,path2,file_path):
    path_new=Path(path1)
    path_old=Path(path2)
    compare_excel(path_new, path_old,file_path)


# Press the to run the script.
if __name__ == "__main__":
    main_gui()
